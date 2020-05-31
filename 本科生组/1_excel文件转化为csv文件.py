"""
    1、将excel文件转化为：以各指数成分股票命名的csv文件，方便后面的pandas以及matplotlib和numpy的操作
"""

# 步骤一：将excel文件转化为csv文件
import os
import csv
from openpyxl import load_workbook

# 打开excel文件
wb = load_workbook('附件：十支股票参数.xlsx')
# 列出文件中所有的表名
sheets = wb.sheetnames

path = 'csv数据集'
if not os.path.exists(path):
    os.mkdir(path)

def create_csvs(title,path):
    name = title
    path = path
    f = open(path + '/' + name + '.csv',mode="a+",newline='',encoding="utf-8-sig")
    csv_write = csv.writer(f)
    csv_write.writerow(['Date', 'Open', 'High', 'Low','Close','Turnover'])
    f.close()

for sheet in sheets:
    create_csvs(sheet,path)
    f = open(path + '/' + sheet + '.csv',mode="a+",newline='',encoding="utf-8-sig")
    csv_write = csv.writer(f)
    table = wb[sheet]
    rows = table.max_row
    cols = table.max_column
    for row in range(5,rows+1):
        data = []
        for col in range(cols+1):
            data.append(str(table.cell(row,col+1).value).replace(r'/','-'))
        csv_write.writerow(data[:6])
    f.close()
