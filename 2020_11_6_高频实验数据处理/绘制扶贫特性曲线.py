import openpyxl
import matplotlib.pyplot as plt
# 找出峰值点
from scipy import signal
import numpy as np
import shapely.geometry as SG

wb = openpyxl.load_workbook('第一次实验数据.xlsx')
sheetnames = wb.sheetnames
sheet = wb['Sheet1']
# 读取excel的行数和列数
rows = sheet.max_row
columns = sheet.max_column
# 需要绘制的曲线就在第一列和第二列
x_values = []
y_values = []
for i in range(2,rows+1):
    for j in range(1,columns+1):
        if j == 1:
            x_values.append(float(sheet.cell(row=i,column=j).value))
        if j == 2:
            y_values.append(float(sheet.cell(row=i,column=j).value))
        else:
            pass
# 使用七次多项式进行拟合
z1 = np.polyfit(x_values,y_values,100)
p1 = np.poly1d(z1)          # 多项式系数
# 打印多项式
# print(p1)
yvals = p1(x_values)
# 求极值
num_peeks = signal.find_peaks(yvals,distance=10)
print(x_values[num_peeks[0][0]])
print(yvals[num_peeks[0]][0])

plt.plot(x_values,y_values,label='幅频特性曲线')
plt.plot(x_values,yvals,'r',label='拟合曲线')
plt.xlabel('输入信号频率（MHz）')
plt.ylabel('输出电压幅值（mv）')
plt.title('幅频特性曲线')
# 使中文可以正常显示出来
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
plt.legend()
plt.savefig('first_fix.png')
plt.show()

# 带宽是峰峰值的2 ** (1/2)处两点间的距离
y_value = float(yvals[num_peeks[0]][0]) * (2 ** (1/2))/2
line = SG.LineString(list(zip(x_values,yvals)))
yline = SG.LineString([(8,y_value),(12,y_value)])
coords = np.array(line.intersection(yline))
print(coords[1][0] - coords[0][0])


# 按下R12之后幅频特性曲线的绘制
x_values = []
y_values = []
for i in range(2,rows+1):
    for j in range(1,columns+1):
        if j == 1:
            x_values.append(float(sheet.cell(row=i,column=j).value))
        if j == 4:
            y_values.append(float(sheet.cell(row=i,column=j).value))
        else:
            pass
# 使用七次多项式进行拟合
z1 = np.polyfit(x_values,y_values,100)
p1 = np.poly1d(z1)          # 多项式系数
# 打印多项式
# print(p1)
yvals = p1(x_values)
# 求极值
num_peeks = signal.find_peaks(yvals,distance=10)
print(x_values[num_peeks[0][0]])
print(yvals[num_peeks[0]][0])

plt.plot(x_values,y_values,label='幅频特性曲线')
plt.plot(x_values,yvals,'r',label='拟合曲线')
plt.xlabel('输入信号频率（MHz）')
plt.ylabel('输出电压幅值（mv）')
plt.title('幅频特性曲线')
# 使中文可以正常显示出来
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
plt.legend()
plt.savefig('second_fix.png')
plt.show()




wb = openpyxl.load_workbook('数据2.xlsx')
sheetnames = wb.sheetnames
sheet = wb['Sheet1']
# 读取excel的行数和列数
rows = sheet.max_row
columns = sheet.max_column
# 需要绘制的曲线就在第一列和第二列
x_values = []
y_values = []
for i in range(2,rows+1):
    for j in range(1,columns+1):
        if j == 1:
            x_values.append(float(sheet.cell(row=i,column=j).value))
        if j == 3:
            y_values.append(float(sheet.cell(row=i,column=j).value))
        else:
            pass
# 使用七次多项式进行拟合
z1 = np.polyfit(x_values,y_values,100)
p1 = np.poly1d(z1)          # 多项式系数
# 打印多项式
# print(p1)
yvals = p1(x_values)
# 求极值
num_peeks = signal.find_peaks(yvals,distance=10)
print(x_values[num_peeks[0][0]])
print(yvals[num_peeks[0]][0])

plt.plot(x_values,y_values,label='幅频特性曲线')
plt.plot(x_values,yvals,'r',label='拟合曲线')
plt.xlabel('输入信号频率（MHz）')
plt.ylabel('输出电压幅值（mv）')
plt.title('幅频特性曲线')
# 使中文可以正常显示出来
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
plt.legend()
plt.savefig('third_fix.png')
plt.show()