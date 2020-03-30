"""
    目标地址：http://www.chinatax.gov.cn/chinatax/n810346/n810825/index.html
    1. 采集
    采集以下栏目：
        增值税、消费税的前两页数据
    每个栏目的：
        标题、发文日期、文号

    2. 保存
        将所有信息保存到`税务局.xlsx`文件，依据栏目名创建数据表，每个栏目的数据保存到对应的数据表
        将所有信息分别保存到`税务局-增值税.json` `税务局-消费税.json`文件
        将所有信息分别保存到`税务局-增值税.csv` `税务局-消费税.csv`文件
        将所有信息分别保存到`shuiwujuzengzhi`数据表与`shuiwujuxiaofei`数据表
"""
import requests
import time
import csv
from openpyxl import Workbook
import json
import pymysql
import pymongo


# 先处理增值税数据
# 增值税41页，消费税9页
def get_VAT_datas(kd, page):
    url = 'http://www.chinatax.gov.cn/sfc/query.ejf?title=fgk&method=json'
    headers = {
        'Cookie': 'JSESSIONID=BE19DBF7221EA38F34FA24BF82D0FE6B; yfx_c_g_u_id_10003701=_ck20033008470614919998579947691; yfx_f_l_v_t_10003701=f_t_1585529226488__r_t_1585529226488__v_t_1585529226488__r_c_0; _Jo0OQK=7FC4DD368D5CFAB62E58537888BFCADCA92BC21F3F266087BA48E60B495A95AB753A3EE58BEFA28AE5FA87FD858E1B97ED965A8D7EA1E83A75E9860124F0B86C06B34275DAD340EB4DD7EF4EAE0C6A2C5647EF4EAE0C6A2C56491A0E8A83004FA3FGJ1Z1Ug==',
        'Host': 'www.chinatax.gov.cn',
        'Origin': 'http://www.chinatax.gov.cn',
        'Referer': 'http://www.chinatax.gov.cn/chinatax/n810346/n810825/index.html',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'
    }
    data = {
        'C2': str(kd),
        'page': str(page),
        'pageSize': '15',
        'qt': '*'
    }
    response = requests.post(url=url, headers=headers, data=data)
    results = response.json()['resultList']
    data = [[] for i in range(len(results))]
    for i in range(len(results)):
        data[i].append(results[i]['dreTitle'])  # 标题dreTitle
        data[i].append(results[i]['myValues']['_FWRQQ'])  # 发文日期docDate
        data[i].append(results[i]['myValues']['DOCNO'])  # 文号DOCNO
        data[i].append(results[i]['myValues']['DRECONTENT'])  # 文章内容DRECONTENT
    return (data)

def save_datas_to_csv(kd, page):
    f = open("税务局-{}".format(kd) + '.csv', mode="a+", newline="", encoding="utf-8-sig")
    csv_write = csv.writer(f)
    csv_write.writerow(['标题', '发文日期', '文号', '文章内容'])
    datas = get_VAT_datas(kd, page)
    for data in datas:
        csv_write.writerow(data)
    f.close()

def save_datas_to_xlsx(kd, page):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '增值税'
    ws2 = wb.create_sheet('sheet2')
    ws2.title = '消费税'
    if kd == '增值税':
        datas = get_VAT_datas(kd, page)
        for i in range(len(datas)):
            ws1.cell(row=i + 1 + 15 * (page - 1), column=1).value = datas[i][0]
            ws1.cell(row=i + 1 + 15 * (page - 1), column=2).value = datas[i][1]
            ws1.cell(row=i + 1 + 15 * (page - 1), column=3).value = datas[i][2]
            ws1.cell(row=i + 1 + 15 * (page - 1), column=4).value = datas[i][3]
    elif kd == '消费税':
        datas = get_VAT_datas(kd, page)
        for i in range(len(datas)):
            ws2.cell(row=i + 1 + 15 * (page - 1), column=1).value = datas[i][0]
            ws2.cell(row=i + 1 + 15 * (page - 1), column=2).value = datas[i][1]
            ws2.cell(row=i + 1 + 15 * (page - 1), column=3).value = datas[i][2]
            ws2.cell(row=i + 1 + 15 * (page - 1), column=4).value = datas[i][3]
    else:
        print("输入的关键词不包含消费税和增值税！")
    wb.save("税务局.xlsx")

def save_datas_to_json(kd,page):
    f = open("税务局-{}.json".format(kd),mode="w",encoding="utf-8")
    datas = get_VAT_datas(kd,page)
    for data in datas:
        f.write(json.dumps(str(data),ensure_ascii=False))
        f.write('\n')
    f.close()

def save_datas_to_mysql(kd,page):
    """
        create database shuiwuju character set utf8;
        use shuiwuju;
        create table zenzhishui(
          title char(100),
          date char(50),
          docno char(100),
          content TEXT(1000)
        )character set utf8;

        create table xiaofeishui(
          title char(100),
          date char(50),
          docno char(100),
          content TEXT(1000)
        )character set utf8;
    """
    connection = pymysql.connect(
        host = 'localhost',
        port = 3306,
        user = 'root',
        password = 'wangzhi20001115',
        database = 'shuiwuju'
    )
    cursor = connection.cursor()
    datas = get_VAT_datas(kd,page)
    if kd == '增值税':
        for data in datas:
            cursor.execute("insert into zenzhishui(title,date,docno,content) values(%s,%s,%s,%s)",data)
    elif kd == '消费税':
        for data in datas:
            cursor.execute("insert into xiaofeishui(title,date,docno,content) values(%s,%s,%s,%s)",data)
    connection.commit()
    cursor.close()
    connection.close()

def save_datas_to_mongodb(kd,page):
    # 连接MongoDB
    client = pymongo.MongoClient(host="localhost",port=27017)
    datas = get_VAT_datas(kd,page)
    for data in datas:
        dictlist = {}
        dictlist['标题'] = data[0]
        dictlist['发文日期'] = data[1]
        dictlist['文号'] = data[2]
        dictlist['文章内容'] = data[3]
        if kd == '消费税':
            db = client.shuiwuju
            collection = db.xiaofeishui
            collection.insert_one(dictlist)
        elif kd == '增值税':
            db = client.shuiwuju
            collection = db.zenzhishui
            collection.insert_one(dictlist)

if __name__ == '__main__':
    for i in range(1,42):
        save_datas_to_csv("增值税",i)
        save_datas_to_xlsx("增值税",i)
        save_datas_to_json("增值税",i)
        save_datas_to_mysql("增值税",i)
        save_datas_to_mongodb("增值税",i)
        print('第{}页数据分别保存到csv、Excel、json、mysql、MongoDB中！'.format(i))
        time.sleep(5)
    for j in range(1,10):
        save_datas_to_csv("消费税",j)
        save_datas_to_xlsx("消费税",j)
        save_datas_to_json("消费税",j)
        save_datas_to_mysql("消费税",j)
        save_datas_to_mongodb("消费税",j)
        print('第{}页数据分别保存到csv、Excel、json、mysql、MongoDB中！'.format(j))
        time.sleep(5)