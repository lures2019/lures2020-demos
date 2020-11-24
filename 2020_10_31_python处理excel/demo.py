import openpyxl

wb = openpyxl.load_workbook('000.xlsx')
# 获取所有的sheet名称————['光明中学', '红星中学', '阳光中学', '灿烂中学']
tables = wb.sheetnames
# 建立列表来分别存储这四个sheet的数据
data_lists = [[] for i in range(len(tables))]
# 查看sheet表的最大行数和最大列数
for table in tables:
    sheet = wb[table]
    rows = sheet.max_row
    columns = sheet.max_column
    for i in range(len(tables)):
        if tables[i] == table:
            # 以行的形式将数据存放到列表中
            for row in sheet.iter_rows(min_row=1,min_col=1,max_col=columns,max_row=rows):
                data_lists[i].append(row)
# 建立新表
new_wb = openpyxl.Workbook()
ws = new_wb.active
ws.title = "新合成表"
for table in tables:
    new_wb.create_sheet(table)
    new_sheet = new_wb[table]
    for i in range(len(tables)):
        if tables[i] == table:
            for q in range(len(data_lists[i])):
                for j in range(len(data_lists[i][q])):
                    new_sheet.cell(row=q+1,column=j+1).value = data_lists[i][q][j].value
                if q == len(data_lists[i]) - 1:
                    # 添加sheet至每个sheet的最大列右边一列，最大行的下面一行
                    new_sheet.cell(row=1,column=j+2).value = table
for i in range(len(tables)):
    m = 1
    n = 1
    # 光明中学无要合并的数据
    if tables[i] == '光明中学':
        pass
    elif tables[i] == '阳光中学':
        for q in range(len(data_lists[i])):
            for j in range(len(data_lists[i][q])):
                # ws.cell(row=q + 1, column=j + 1).value = data_lists[i][q][j].value
                if (q >= 14 and q <= 21) and j <= 9:
                    ws.cell(row = m, column = n).value = data_lists[i][q][j].value
                    n += 1
                    if (n % 9 == 0):
                        m += 1
                        n -= 8
        m += 1
        if tables[i + 1] == '灿烂中学':
            for q in range(len(data_lists[i+1])):
                for j in range(len(data_lists[i+1][q])):
                    # ws.cell(row=q + 1, column=j + 1).value = data_lists[i][q][j].value
                    if (q >= 14 and q <= 21) and j <= 9:
                        ws.cell(row=m, column=n).value = data_lists[i+1][q][j].value
                        n += 1
                        if (n % 9 == 0):
                            m += 1
                            n -= 8
        m += 1
        if tables[i - 1] == '红星中学':
            for q in range(len(data_lists[i - 1])):
                for j in range(len(data_lists[i - 1][q])):
                    # ws.cell(row=q + 1, column=j + 1).value = data_lists[i][q][j].value
                    if (q >= 11 and q <= 17) and j <= 9:
                        ws.cell(row=m, column=n).value = data_lists[i - 1][q][j].value
                        n += 1
                        if (n % 9 == 0):
                            m += 1
                            n -= 8
new_wb.save('处理后excel.xlsx')
