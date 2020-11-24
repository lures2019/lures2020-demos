"""
    将北京、大连、天津的日期、AQI指数、质量等级、PM2.5、PM10、SO2、NO2、CO、O3信息爬取出来，时间为2018年1月1日到现在。
    代码思路：
        首先，观察需要爬取的三个城市的数据的url的区别
        其次，具体分析一个城市的某个月份的数据的解析方式
        最后，将爬取到的数据全部写入excel文件中
"""

import requests
import parsel
from openpyxl import Workbook
from openpyxl import  load_workbook
import time

def save_data_to_excel(url,city,number):
    url = url
    city = city
    number = number
    headers = {
        'Cookie': 'ASP.NET_SessionId=qeiero45wltlpr2xcxxfafa3; bdshare_firstime=1598598193109; Hm_lvt_f48cedd6a69101030e93d4ef60f48fd0=1598598195; __51cke__=; __gads=ID=ba5a0d97129a6d82:T=1598598195:S=ALNI_Mbiyr6oDP--gGwYxTeU7zDSB8v2nw; Hm_lpvt_f48cedd6a69101030e93d4ef60f48fd0=1598600488; __tins__4560568=%7B%22sid%22%3A%201598600488153%2C%20%22vd%22%3A%201%2C%20%22expires%22%3A%201598602288153%7D; __51laig__=6',
        'Host': 'www.tianqihoubao.com',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36'
    }
    response = requests.get(url=url, headers=headers)
    response.encoding = response.apparent_encoding
    html = response.text
    select = parsel.Selector(html)
    # 需要爬取的几个数据项的名称
    titles = select.xpath('//div[@class="api_month_list"]/table[@class="b"]/tr[1]/td/b/text()').getall()
    informations = select.xpath('//div[@class="api_month_list"]/table[@class="b"]/tr/td/text()').getall()
    # 将所有的数据存放到列表中
    results = []
    for i in range(10, len(informations)):
        results.append(informations[i].replace('\n', '').strip())

    # 下面准备将数据写入excel中
    work = load_workbook('天气.xlsx')
    if city == 'beijing':
        g = work.sheetnames[0]
    elif city == 'tianjin':
        g = work.sheetnames[1]
    else:
        g = work.sheetnames[2]
    sheet = work[g]
    for i in range(len(titles)):
        sheet.cell(row=1+number*31, column=i + 1).value = titles[i]
    lines = [[] for i in range(int(len(results) / len(titles)))]
    n = len(titles)  # 大列表中几个数据组成一个小列表
    results = [results[i:i + n] for i in range(0, len(results), n)]
    for j in range(len(lines)):
        for x in range(len(titles)):
            sheet.cell(row=j + 2+number*31, column=x + 1).value = results[j][x]

    work.save('天气.xlsx')


if __name__ == '__main__':
    # 创建Excel及对应的表明
    wb = Workbook()
    ws = wb.active
    ws.title = '北京历史天气'
    wb.create_sheet('天津历史天气')
    wb.create_sheet('大连历史天气')
    wb.save('天气.xlsx')

    cities = ['beijing','tianjin','dalian']
    years = ['2018','2019','2020']
    months = ['01','02','03','04','05','06','07','08','09','10','11','12']
    for city in cities:
        number = 0
        for year in years:
            if year != '2020':
                for month in months:
                    url = 'http://www.tianqihoubao.com/aqi/{}-{}{}.html'.format(city,year,month)
                    save_data_to_excel(url,city,number)
                    number += 1
                    time.sleep(3)
                print('{}市{}年历史天气保存完毕！'.format(city,year))
            else:
                for month in ['01','02','03','04','05','06','07']:
                    url = 'http://www.tianqihoubao.com/aqi/{}-{}{}.html'.format(city,year,month)
                    save_data_to_excel(url,city,number)
                    number += 1
                    time.sleep(3)
                print('{}市{}年历史天气保存完毕！'.format(city,year))