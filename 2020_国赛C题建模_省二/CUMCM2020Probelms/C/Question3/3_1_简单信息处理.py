from openpyxl import load_workbook
from openpyxl import  Workbook
import csv
import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn import metrics
from sklearn.externals import  joblib


work = load_workbook('企业分类.xlsx')
sheet = work['Sheet1']
# 分别求当前sheet表的行数和列数
rows = sheet.max_row
columns = sheet.max_column

# 创建空字典来计算各行各业分类的数据
my_dict = {}
# 现在开始使用循环遍历sheet
for i in range(1,rows+1):
    key_name = sheet.cell(row=i,column=3).value
    if key_name == None:
        pass
    else:
        key_name = key_name.strip()
        if key_name not in my_dict:
            my_dict[key_name] = 1
        else:
            my_dict[key_name] += 1

# 设置变量存储企业的名称以及数目
names = []
numbers = []
for key,value in my_dict.items():
    names.append(key)
    numbers.append(value)
# 创建列表来存储每个行业的所有公司的代号
lists = [[] for i in range(len(names))]
for i in range(len(names)):
    for j in range(1,rows+1):
        key_now = sheet.cell(row=j,column=3).value
        value_now = sheet.cell(row=j,column=1).value
        if key_now == names[i]:
            lists[i].append(value_now)

f = open('../Question2/Task1/8_处理后信息.csv',mode='r',encoding='utf-8')
csv_reader = csv.reader(f)
# 创建一个存储数值的字典
money = {}
for row in csv_reader:
    if row[1] == '企业代号':
        pass
    else:
        money[row[1]] = row[12]

# 现在开始进行匹配，将对应编号的净利润输入前面的各小列表中
# 先创建一个大列表，存储各值
last_money = [[] for i in range(len(names))]
for i in range(len(lists)):
    for j in lists[i]:
        last_money[i].append(money[j])

# 现在开始计算每个列表里数据的方差
# 先创建空列表存储平均值
average = []
for i in range(len(last_money)):
    total = 0
    for m in last_money[i]:
        total += float(m)
    average.append(total/numbers[i])
# 现在开始求方差
variance = []
for i in range(len(last_money)):
    total_now = 0
    for x in last_money[i]:
        total_now += (float(x) - average[i]) ** 2
    variance.append(total_now/numbers[i])
# 现在将行业名称和方差对应起来
new_dict = {}
for k in range(len(names)):
    new_dict[names[k]] = variance[k]

wb = Workbook()
ws = wb.active
ws.title = '行业信息'

for i in range(1,len(names)+1):
    for j in range(1,len(lists[i-1])+1):
        ws.cell(row=j,column=i).value = lists[i-1][j-1]

wb.save('信息.xlsx')