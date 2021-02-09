from openpyxl import load_workbook
import os
import csv
import matplotlib.pyplot as plt
from mpl_toolkits.basemap import Basemap
import pandas as pd
import numpy as np


path = "datas"
# 1、创建文件夹用于保存处理后的数据以及可视化图片
if not os.path.exists(path):
    os.mkdir(path)

# 2、将.xlsx换成.csv格式
def exchange_xlsx_to_csv(path):
    wb = load_workbook("../2021_MCM_Problem_C_Data/2021MCMProblemC_DataSet.xlsx")
    sheet = wb['Sheet1']
    # 得到行数和列数，方便读取每一个单元格的值
    nrows = sheet.max_row
    ncols = sheet.max_column
    # 将xlsx格式整体改成csv格式
    new_files = []
    for i in range(1, nrows + 1):
        rows = []
        for j in range(1, ncols + 1):
            value = sheet.cell(row=i, column=j).value
            rows.append(value)
        new_files.append(rows)
    # 将数据写入csv
    f = open(path + '/' + '2021MCMProblemC_DataSet.csv', mode="w+", newline="", encoding="utf-8-sig")
    csv_write = csv.writer(f)
    for i in range(len(new_files)):
        if i >= 1:
            new_files[i][-2] = float(new_files[i][-2])
            new_files[i][-1] = float(new_files[i][-1])
        csv_write.writerow(new_files[i])
    f.close()

# 3、读取csv文件，进行划分数据集
def divide_several_csv_files(path):
    f = open("datas/2021MCMProblemC_DataSet.csv",mode='r',encoding="utf-8")
    csv_reader = csv.reader(f)
    rows = [i for i in csv_reader]
    plt.rcParams['font.sans-serif'] = 'SimHei'
    plt.rcParams['axes.unicode_minus'] = False
    latitudes = [eval(row[-2]) for row in rows[1:]]
    longitudes = [eval(row[-1]) for row in rows[1:]]
    fig, ax = plt.subplots(figsize=(9, 6))
    ax.scatter(latitudes, longitudes, alpha=0.3)
    ax.set_xlabel("latitude")
    ax.set_ylabel("longitude")
    plt.savefig(path + '/' + "2021MCMProblemC_DataSet.png")
    plt.show()
    # 需要的数据是    Lab Status，构造一个字典来统计这个分类情况以及各自的数据量
    lab_status_dict = {}
    for row in rows[1:]:
        key = row[3]
        if key not in lab_status_dict:
            lab_status_dict[key] = 1
        else:
            lab_status_dict[key] += 1
    keys = list(lab_status_dict.keys())
    print("2021MCMProblemC_DataSet.csv数据集中lab_status可以分为：{}".format(keys))
    print("对应的数据量分别是：{}".format(list(lab_status_dict.values())))
    files = [[] for i in range(len(lab_status_dict))]
    for i in range(len(files)):
        for row in rows[1:]:
            if row[3] == keys[i]:
                files[i].append(row)
    # 写入各自的csv文件
    for i in range(len(files)):
        f = open(path + '/' +"{}.csv".format(keys[i]),mode="w+",newline="",encoding="utf-8-sig")
        csv_write = csv.writer(f)
        for row in files[i]:
            csv_write.writerow(row)
        f.close()

# 4、根据经纬度坐标进行可视化(散点图)
def paint_scatter_plot(path):
    filenames = ['Negative ID.csv', 'Positive ID.csv', 'Unprocessed.csv', 'Unverified.csv']
    for i in range(len(filenames)):
        plt.rcParams['font.sans-serif'] = 'SimHei'
        plt.rcParams['axes.unicode_minus'] = False
        f = open("datas/{}".format(filenames[i]),mode="r",encoding="utf-8")
        csv_reader = csv.reader(f)
        rows = [i for i in csv_reader]
        latitudes = [eval(row[-2]) for row in rows]
        longitudes = [eval(row[-1]) for row in rows]
        fig, ax = plt.subplots(figsize=(9, 6))
        ax.scatter(latitudes,longitudes,alpha=0.3)
        ax.set_xlabel("latitude")
        ax.set_ylabel("longitude")
        name = filenames[i].split('.')[0]
        plt.savefig(path + '/' +"{}.png".format(name))
        plt.show()

# 5、根据经纬度进行地图可视化
def paint_map(path):
    # 绘制基础地图，选择绘制的区域，因为是绘制美国地图，故选取如下经纬度，lat_0和lon_0是地图中心的维度和经度
    map = Basemap(projection='stere', lat_0=90, lon_0=-105,llcrnrlat=23.41, urcrnrlat=45.44,llcrnrlon=-118.67, urcrnrlon=-64.52,rsphere=6371200., resolution='l', area_thresh=10000)
    map.drawmapboundary()  # 绘制边界
    map.drawstates()  # 绘制州
    map.drawcoastlines()  # 绘制海岸线
    map.drawcountries()  # 绘制国家
    map.drawcounties()  # 绘制县
    parallels = np.arange(0., 90, 10.)
    map.drawparallels(parallels, labels=[1, 0, 0, 0], fontsize=10)  # 绘制纬线
    meridians = np.arange(-110., -60., 10.)
    map.drawmeridians(meridians, labels=[0, 0, 0, 1], fontsize=10)  # 绘制经线
    posi = pd.read_csv("datas/2021MCMProblemC_DataSet.csv")

    lat = np.array(posi["Longitude"])
    lon = np.array(posi["Latitude"])
    x, y = map(lon, lat)
    map.scatter(x, y,c="g")  # 也可以使用Basemap的methord本身的scatter
    plt.title('America')
    plt.savefig(path + '/' + 'map.png')
    plt.show()

if __name__ == "__main__":
    exchange_xlsx_to_csv(path)
    divide_several_csv_files(path)
    paint_scatter_plot(path)
    paint_map(path)