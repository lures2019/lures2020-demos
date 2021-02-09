import csv
from openpyxl import load_workbook
import os
from shutil import move

# 1、将xlsx格式文件转换成csv文件，方便后续操作
def exchange_xlsx_to_csv():
    wb = load_workbook("../2021_MCM_Problem_C_Data/2021MCM_ProblemC_ Images_by_GlobalID.xlsx", read_only=True)
    sheet = wb['AttachmentRelateTable']
    rows = sheet.max_row
    cols = sheet.max_column
    f = open("datas/2021MCM_ProblemC_ Images_by_GlobalID.csv", mode='w+', newline="", encoding="utf-8-sig")
    csv_write = csv.writer(f)
    for i in range(1, rows + 1):
        row = []
        for j in range(1, cols):
            value = sheet.cell(row=i, column=j).value
            row.append(value)
        csv_write.writerow(row)
    f.close()

# 2、对比两个csv文件
def compare_two_csv():
    f1 = open("datas/2021MCMProblemC_DataSet.csv",mode="r",encoding="utf-8")
    csv_reader = csv.reader(f1)
    rows = [i for i in csv_reader]
    labs = [row[3] for row in rows[1:]]
    ids = [row[0] for row in rows[1:]]
    f2 = open("datas/2021MCM_ProblemC_ Images_by_GlobalID.csv",mode="r",encoding="utf-8")
    csv_reader = csv.reader(f2)
    rows = [i for i in csv_reader]
    fp = open("datas/question2_1_add_status.csv",mode="w+",newline="",encoding="utf-8-sig")
    csv_write = csv.writer(fp)
    rows[0].append("lab_status")
    csv_write.writerow(rows[0])
    for row in rows[1:]:
        for i in range(len(ids)):
            if row[1] == ids[i]:
                row.append(labs[i])
                break
        csv_write.writerow(row)
    fp.close()

# 3、根据得到的csv文件筛选出已经检测出是positive和negative的那些信息
def choose_images_to_train():
    f = open("datas/question2_1_add_status.csv",mode="r",encoding="utf-8")
    csv_reader = csv.reader(f)
    rows = [i for i in csv_reader]
    fp = open("datas/question2_2_choose_images_to_train.csv",mode="w+",newline="",encoding="utf-8-sig")
    csv_write = csv.writer(fp)
    csv_write.writerow(rows[0])
    # 获取图片格式有{'jpg': 3032, 'png': 79, 'quicktime': 76, 'vnd.openxmlformats-officedocument.wordprocessingml.document': 3, 'mp4': 9, 'x-zip-compressed': 3, 'pdf': 5, 'octet-stream': 1}
    # 只选择jpg和png，都转换为jpg格式
    for row in rows[1:]:
        form = row[-2].split("/")[-1]
        new_row = []
        if (row[-1] == "Negative ID") or (row[-1] == "Positive ID") and (form == "jpg" or form == "png"):
            form0 = row[0].split(".")[0] + ".jpg"
            form1 = row[1]
            form2 = row[-2].split("/")[0] + ".jpg"
            form3 = row[-1]
            new_row.append(form0)
            new_row.append(form1)
            new_row.append(form2)
            new_row.append(form3)
            csv_write.writerow(new_row)
    fp.close()

# 4、筛选图片文件，将其他资源信息单独放入一个文件夹内
def put_other_resources_into_folder():
    path = "../2021MCM_ProblemC_Files/other_resources"
    if not os.path.exists(path):
        os.mkdir(path)
    path_make = "../2021MCM_ProblemC_Files"
    files = os.listdir(path_make)
    for file in files:
        if file == "other_resources":
            pass
        else:
            form = file.split(".")[-1]
            name = file.split(".")[0]
            if (form.lower() == "jpg") or (form.lower() == "png") or (form == name):
                if (form.lower() == "png") or (form == name):
                    newname = "{}.jpg".format(name)
                    # 注意代码的路径和图片路径不一致
                    src_path = path_make + "/" + file
                    dest_path = path_make + "/" +newname
                    os.rename(src_path,dest_path)
            else:
                src_path = "../2021MCM_ProblemC_Files/{}".format(file)
                move(src_path,path)

if __name__ == '__main__':
    exchange_xlsx_to_csv()
    compare_two_csv()
    choose_images_to_train()
    # put_other_resources_into_folder()